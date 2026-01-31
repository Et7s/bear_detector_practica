import os
import json
from datetime import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors


from config import RESULT_FOLDER

def setup_excel_styles():
    styles = {
        'header_font': Font(name='Arial', size=12, bold=True, color='FFFFFF'),
        'header_fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
        'subheader_font': Font(name='Arial', size=11, bold=True, color='000000'),
        'subheader_fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
        'data_font': Font(name='Arial', size=10),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'center_alignment': Alignment(horizontal='center', vertical='center'),
        'left_alignment': Alignment(horizontal='left', vertical='center'),
        'right_alignment': Alignment(horizontal='right', vertical='center')
    }
    
    # Форматы для чисел
    styles['percent_format'] = '0.00%'
    styles['date_format'] = 'DD.MM.YYYY HH:MM:SS'
    
    return styles

def generate_excel_report(history_data):
    if not history_data:
        return create_empty_report()
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"bear_detection_report_{timestamp}.xlsx"
    excel_path = os.path.join(RESULT_FOLDER, excel_filename)
    
    # Создаем Excel workbook
    wb = Workbook()
    styles = setup_excel_styles()
    
    # Удаляем дефолтный лист
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 1. Лист "Сводка" (Summary)
    create_summary_sheet(wb, styles, history_data)
    
    # 2. Лист "Детализация" (Details)
    create_details_sheet(wb, styles, history_data)
    
    # Сохраняем Excel файл
    wb.save(excel_path)
    
    print(f"✅ Excel отчет создан: {excel_path}")
    return excel_path

def create_empty_report():
    """Создание пустого отчета когда нет данных"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"empty_report_{timestamp}.xlsx"
    excel_path = os.path.join(RESULT_FOLDER, excel_filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    
    ws['A1'] = "Отчет по детекции медведей"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:E1')
    
    ws['A3'] = "В системе пока нет данных для отчета"
    ws['A4'] = "Загрузите изображения с медведями для начала работы"
    
    for column in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[column].width = 20
    
    wb.save(excel_path)
    return excel_path

def create_summary_sheet(wb, styles, history_data):
    ws = wb.create_sheet(title="Сводка")
    
    # Заголовок отчета
    ws['A1'] = "ОТЧЕТ ПО ДЕТЕКЦИИ МЕДВЕДЕЙ"
    ws['A1'].font = Font(size=16, bold=True, color='366092')
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = styles['center_alignment']
    
    # Информация о отчете
    ws['A3'] = "Дата генерации:"
    ws['B3'] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    ws['A4'] = "Период данных:"
    
    if history_data:
        dates = [datetime.fromisoformat(h['timestamp']) for h in history_data]
        min_date = min(dates).strftime("%d.%m.%Y")
        max_date = max(dates).strftime("%d.%m.%Y")
        ws['B4'] = f"{min_date} - {max_date}"
    else:
        ws['B4'] = "Нет данных"
    
    # Основная статистика
    row = 6
    ws[f'A{row}'] = "ОСНОВНЫЕ ПОКАЗАТЕЛИ"
    ws[f'A{row}'].font = styles['subheader_font']
    ws[f'A{row}'].fill = styles['subheader_fill']
    ws.merge_cells(f'A{row}:G{row}')
    
    row += 2
    
    # Импортируем функцию calculate_summary_statistics из history_manager
    from utils.history_manager import calculate_summary_statistics
    summary_data = calculate_summary_statistics(history_data)
    
    ws[f'A{row}'] = "Всего запросов"
    ws[f'B{row}'] = summary_data['total_requests']
    row += 1
    
    ws[f'A{row}'] = "Всего обнаружено медведей"
    ws[f'B{row}'] = summary_data['total_bears']
    row += 1
    
    ws[f'A{row}'] = "Средняя уверенность"
    ws[f'B{row}'] = summary_data['avg_confidence']
    ws[f'B{row}'].number_format = styles['percent_format']
    row += 1
    
    ws[f'A{row}'] = "Максимальная уверенность"
    ws[f'B{row}'] = summary_data['max_confidence']
    ws[f'B{row}'].number_format = styles['percent_format']
    row += 1
    
    ws[f'A{row}'] = "Минимальная уверенность"
    ws[f'B{row}'] = summary_data['min_confidence']
    ws[f'B{row}'].number_format = styles['percent_format']
    row += 1
    
    ws[f'A{row}'] = "Медведей на запрос (средн.)"
    ws[f'B{row}'] = summary_data['bears_per_request']
    ws[f'B{row}'].number_format = '0.00'
    row += 2
    
    # Статистика по дням
    ws[f'A{row}'] = "СТАТИСТИКА ПО ДНЯМ"
    ws[f'A{row}'].font = styles['subheader_font']
    ws[f'A{row}'].fill = styles['subheader_fill']
    ws.merge_cells(f'A{row}:G{row}')
    row += 2
    
    daily_stats = summary_data['daily_stats']
    
    # Заголовки таблицы
    headers = ['Дата', 'Запросов', 'Медведей', 'Ср. уверенность', 'Макс. уверенность']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['center_alignment']
        cell.border = styles['border']
    
    row += 1
    
    # Данные по дням
    for day_stat in daily_stats:
        ws.cell(row=row, column=1, value=day_stat['date']).border = styles['border']
        ws.cell(row=row, column=2, value=day_stat['count']).border = styles['border']
        ws.cell(row=row, column=3, value=day_stat['bears']).border = styles['border']
        ws.cell(row=row, column=4, value=day_stat['avg_confidence']).border = styles['border']
        ws.cell(row=row, column=4).number_format = styles['percent_format']
        ws.cell(row=row, column=5, value=day_stat['max_confidence']).border = styles['border']
        ws.cell(row=row, column=5).number_format = styles['percent_format']
        row += 1
    
    # Автоподбор ширины колонок
    for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[column].width = 20

def create_details_sheet(wb, styles, history_data):
    ws = wb.create_sheet(title="Детализация")
    
    row = 1
    ws[f'A{row}'] = "ДЕТАЛИЗАЦИЯ ВСЕХ ЗАПРОСОВ"
    ws[f'A{row}'].font = Font(size=14, bold=True, color='366092')
    ws.merge_cells(f'A{row}:J{row}')  # Было L (12 колонок), стало J (10 колонок)
    ws[f'A{row}'].alignment = styles['center_alignment']
    
    row += 2
    
    # Заголовки таблицы (убрали колонки с площадями)
    headers = [
        '№', 'Дата и время', 'Тип файла', 'Медведей', 
        'Уверенность (ср.)', 'Уверенность (макс.)', 'Уверенность (мин.)',
        'Файл', 'ID события', 'Статус'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['center_alignment']
        cell.border = styles['border']
    
    row += 1
    
    # Данные запросов
    for idx, item in enumerate(history_data, 1):
        # Расчет метрик (только уверенности, без площадей)
        detections = item['detections']
        
        if detections:
            confidences = [d['confidence'] for d in detections]
            avg_confidence = np.mean(confidences)
            max_confidence = max(confidences)
            min_confidence = min(confidences)
        else:
            avg_confidence = max_confidence = min_confidence = 0
        
        # Определение типа файла
        filename = item.get('original_image', '')
        if filename.lower().endswith(('.mp4', '.avi', '.mov')):
            file_type = 'Видео'
        else:
            file_type = 'Изображение'
        
        # Определение статуса
        status = "Успешно" if detections else "Не обнаружено"
        
        # Заполнение строки (убрали колонки с площадями)
        ws.cell(row=row, column=1, value=idx).border = styles['border']
        ws.cell(row=row, column=2, value=item['timestamp']).border = styles['border']
        ws.cell(row=row, column=3, value=file_type).border = styles['border']
        ws.cell(row=row, column=4, value=item['bear_count']).border = styles['border']
        ws.cell(row=row, column=5, value=avg_confidence).border = styles['border']
        ws.cell(row=row, column=5).number_format = styles['percent_format']
        ws.cell(row=row, column=6, value=max_confidence).border = styles['border']
        ws.cell(row=row, column=6).number_format = styles['percent_format']
        ws.cell(row=row, column=7, value=min_confidence).border = styles['border']
        ws.cell(row=row, column=7).number_format = styles['percent_format']
        ws.cell(row=row, column=8, value=os.path.basename(filename)).border = styles['border']
        ws.cell(row=row, column=9, value=item['id']).border = styles['border']
        ws.cell(row=row, column=10, value=status).border = styles['border']
        
        row += 1
    
    # Автоподбор ширины колонок (обновили ширины)
    column_widths = {
        'A': 5,    # №
        'B': 20,   # Дата и время
        'C': 12,   # Тип файла
        'D': 10,   # Медведей
        'E': 15,   # Уверенность (ср.)
        'F': 15,   # Уверенность (макс.)
        'G': 15,   # Уверенность (мин.)
        'H': 25,   # Файл
        'I': 15,   # ID события
        'J': 15    # Статус
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

def generate_json_report(history_data):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"bear_detection_report_{timestamp}.json"
    file_path = os.path.join(RESULT_FOLDER, filename)

    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(history_data, f, ensure_ascii=False, indent=2)

    print(f"✅ JSON отчет создан: {file_path}")
    return file_path

def register_pdf_fonts():
    font_path = os.path.join('fonts', 'DejaVuSans.ttf')
    pdfmetrics.registerFont(TTFont('DejaVu', font_path))



def generate_pdf_report(history_data):
    register_pdf_fonts()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"bear_detection_report_{timestamp}.pdf"
    path = os.path.join(RESULT_FOLDER, filename)

    doc = SimpleDocTemplate(path, pagesize=A4)
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name='TitleDeja',
        fontName='DejaVu',
        fontSize=16,
        alignment=1
    ))

    styles.add(ParagraphStyle(
        name='TextDeja',
        fontName='DejaVu',
        fontSize=10
    ))

    elements = []

    elements.append(Paragraph(
        "Отчёт по детекции медведей",
        styles['TitleDeja']
    ))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(
        f"Дата генерации: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}",
        styles['TextDeja']
    ))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(
        f"Всего запросов: {len(history_data)}",
        styles['TextDeja']
    ))

    total_bears = sum(item['bear_count'] for item in history_data)
    elements.append(Paragraph(
        f"Всего обнаружено медведей: {total_bears}",
        styles['TextDeja']
    ))

    elements.append(Spacer(1, 20))

    table_data = [
        ['Дата', 'Медведей', 'Время обработки']
    ]

    for item in history_data[-15:]:
        table_data.append([
            item['timestamp'],
            str(item['bear_count']),
            f"{item['processing_time']:.2f} сек"
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'DejaVu'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER')
    ]))

    elements.append(table)

    doc.build(elements)
    return path
